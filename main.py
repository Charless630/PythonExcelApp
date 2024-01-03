# Importing tkinter module
import tkinter as tk
from tkinter import ttk
import openpyxl as xl

# Function to load the data
def load_data():
    path = "people.xlsx"
    workbook = xl.load_workbook(path)
    sheet = workbook.active
#   print(sheet.max_row)
    list_values = list(sheet.values)
    print(list_values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)
#    treeview.heading("Name", text="Name")  
    for value_tuple in list_values[1:]:
        treeview.insert("", tk.END, values=value_tuple)
    
# Function to insert the data
def insert_row():
    name = name_entry.get()
    age = int(age_spinbox.get())
    subscription_status = status_combobox.get()
    employment_status = "Employed" if a.get() else "Unemployed"
    
    print(name, age, subscription_status, employment_status)
    
    # Insert row into Excel Sheet
    path = "people.xlsx"
    workbook = xl.load_workbook(path)
    sheet = workbook.active
    row_values = [name, age, subscription_status, employment_status]
    sheet.append(row_values)
    workbook.save(path)
    
    # Insert row into Treeview
    treeview.insert("", tk.END, values=row_values)
    
    # Clear the values
    name_entry.delete(0, 'end')
    name_entry.insert(0, "Name")
    age_spinbox.delete(0, 'end')
    age_spinbox.insert(0, "Age")
    status_combobox.set(combo_list[0])
    checkbutton.state(['!selected'])
    

# Function to toggle between the themes
def toggle_mode():
    if mode_switch.instate(['selected']):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

# Creating root - Actual window
root = tk.Tk()

# Calling the styles
style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

combo_list = ["Subscribed", "Not Subscribed", "Other"]

# Enable the screen to be responsive and resizable
frame = ttk.Frame(root)

# Centering the widgets
frame.pack()

# Creating a label
widget_frame = ttk.LabelFrame(frame, text="Insert Row")
widget_frame.grid(row=0, column=0, padx=20, pady=10)

# Creating an Entry widget
name_entry = ttk.Entry(widget_frame)
name_entry.insert(0, "Name")

# Binding the event to the widget
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

# Creating the Spinbox widget
age_spinbox = ttk.Spinbox(widget_frame, from_=18, to=100)
age_spinbox.insert(0, "Age")

# Binding the event to the Spinbox widget
age_spinbox.bind("<FocusIn>", lambda e: age_spinbox.delete('0', 'end'))
age_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

# Creating the Combobox widget
status_combobox = ttk.Combobox(widget_frame, values=combo_list)
status_combobox.current(0)
status_combobox.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

# Creating the Checkbutton widget
a = tk.BooleanVar()
checkbutton = ttk.Checkbutton(widget_frame, text="Employed", variable=a)
checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

# Creating the Button widget
button = ttk.Button(widget_frame, text="Insert", command=insert_row)
button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

# Creating the Separator widget
separator = ttk.Separator(widget_frame)
separator.grid(row=5, column=0, padx=10, pady=10, sticky="ew")

# Creating the Switch widget
mode_switch = ttk.Checkbutton(widget_frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

# Creating the Treeview widget
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady= 10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

# Creating the Treeview widget
cols = ("Name", "Age", "Subscription", "Employment")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=13)
treeview.column("Name", width=100)
treeview.column("Age", width=50)
treeview.column("Subscription", width=100)
treeview.column("Employment", width=100)
treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()



# Event loop to launch the window
root.mainloop()

